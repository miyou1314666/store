from __future__ import annotations

import csv
import io
import json
import math
import mimetypes
import os
import re
import socket
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote, unquote

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"
GENERATED = ROOT / "generated"
HOST = os.environ.get("FRAME_PARTS_TOOL_HOST", "0.0.0.0")
PORT = int(os.environ.get("PORT") or os.environ.get("FRAME_PARTS_TOOL_PORT", "5508"))

SAP_COLUMNS = ["物料", "物料描述", "需求日期", "需求量", "供应商"]
SRM_COLUMNS = ["计划年份", "计划月份", "供应商编号", "零件编码", "零件名称", "N+1总数"]


def normalize_code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\s+", "", text).upper()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    text = str(value).replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def next_month_label(year: int, month: int) -> str:
    if month >= 12:
        return f"{year + 1}-01"
    return f"{year}-{month + 1:02d}"


def month_key(value: Any) -> str:
    if isinstance(value, datetime):
        return f"{value.year}-{value.month:02d}"
    if isinstance(value, date):
        return f"{value.year}-{value.month:02d}"
    if isinstance(value, (int, float)) and value > 0:
        try:
            dt = from_excel(value)
            return f"{dt.year}-{dt.month:02d}"
        except Exception:
            return ""
    text = clean_text(value)
    if not text:
        return ""
    match = re.search(r"(20\d{2})\D+([01]?\d)", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return f"{year}-{month:02d}"
    match = re.search(r"(20\d{2})([01]\d)", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return f"{year}-{month:02d}"
    return ""


def item_type(code: str, name: str) -> str:
    has_t = "T" in code.upper()
    if "横梁" in str(name) and has_t:
        return "横梁"
    if not has_t:
        return "小件"
    return "排除"


def find_header_indexes(rows: Iterable[tuple[Any, ...]], required: list[str], label: str) -> tuple[dict[str, int], list[tuple[Any, ...]]]:
    cached_rows: list[tuple[Any, ...]] = []
    for row in rows:
        cached_rows.append(row)
        header = {clean_text(cell): idx for idx, cell in enumerate(row)}
        missing = [col for col in required if col not in header]
        if not missing:
            return header, cached_rows
        if len(cached_rows) >= 30:
            raise ValueError(f"{label}缺少必要字段：" + "、".join(missing))
    raise ValueError(f"{label}为空，未读取到数据。")


def iter_excel_records(file_bytes: bytes, filename: str, required: list[str], label: str) -> Iterable[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            row_iter = sheet.iter_rows(values_only=True)
            headers, cached_rows = find_header_indexes(row_iter, required, label)
            for row in cached_rows[1:]:
                yield {col: row[headers[col]] if headers[col] < len(row) else None for col in required}
            for row in row_iter:
                yield {col: row[headers[col]] if headers[col] < len(row) else None for col in required}
        finally:
            workbook.close()
        return

    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("读取 .xls 文件需要安装 xlrd 依赖。") from exc
        book = xlrd.open_workbook(file_contents=file_bytes, on_demand=True)
        try:
            sheet = book.sheet_by_index(0)
            rows = (tuple(sheet.row_values(i)) for i in range(sheet.nrows))
            headers, cached_rows = find_header_indexes(rows, required, label)
            for row in cached_rows[1:]:
                yield {col: row[headers[col]] if headers[col] < len(row) else None for col in required}
            for row in rows:
                yield {col: row[headers[col]] if headers[col] < len(row) else None for col in required}
        finally:
            book.release_resources()
        return

    text: str
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("gbk", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ValueError(f"{label}为空，未读取到数据。")
    missing = [col for col in required if col not in reader.fieldnames]
    if missing:
        raise ValueError(f"{label}缺少必要字段：" + "、".join(missing))
    for row in reader:
        yield {col: row.get(col) for col in required}


def collect_sap(file_bytes: bytes, filename: str) -> tuple[dict[str, dict[str, Any]], list[str], int]:
    info: dict[str, dict[str, Any]] = {}
    month_usage: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    months_seen: set[str] = set()
    row_count = 0

    for row in iter_excel_records(file_bytes, filename, SAP_COLUMNS, "SAP表"):
        row_count += 1
        code = normalize_code(row["物料"])
        month = month_key(row["需求日期"])
        if not code or not month:
            continue

        months_seen.add(month)
        month_usage[code][month] += to_number(row["需求量"])
        if code not in info:
            info[code] = {
                "图号": code,
                "物料名称": clean_text(row["物料描述"]),
                "供应商代码": clean_text(row["供应商"]),
            }
        elif not info[code]["供应商代码"]:
            info[code]["供应商代码"] = clean_text(row["供应商"])

    recent_months = sorted(months_seen)[-3:]
    if len(recent_months) < 3:
        raise ValueError("SAP表中可识别的月份少于3个月，请检查需求日期字段。")

    for code, usage in month_usage.items():
        record = info.setdefault(code, {"图号": code, "物料名称": "", "供应商代码": ""})
        for month in recent_months:
            record[month] = usage.get(month, 0.0)
        record["前三个月月均用量"] = sum(record[month] for month in recent_months) / 3
        record["物料类型"] = item_type(code, record["物料名称"])

    return info, recent_months, row_count


def collect_srm(file_bytes: bytes, filename: str) -> tuple[dict[str, dict[str, Any]], str, int]:
    summary: dict[str, dict[str, Any]] = {}
    year_counter: Counter[int] = Counter()
    month_counter: Counter[int] = Counter()
    row_count = 0

    for row in iter_excel_records(file_bytes, filename, SRM_COLUMNS, "SRM表"):
        row_count += 1
        code = normalize_code(row["零件编码"])
        if not code:
            continue

        demand = to_number(row["N+1总数"])
        supplier = clean_text(row["供应商编号"])
        record = summary.setdefault(code, {"下月预测需求": 0.0, "SRM供应商代码": supplier})
        record["下月预测需求"] += demand
        if not record["SRM供应商代码"] and supplier:
            record["SRM供应商代码"] = supplier

        year = int(to_number(row["计划年份"]))
        month = int(to_number(row["计划月份"]))
        if year > 0:
            year_counter[year] += 1
        if 1 <= month <= 12:
            month_counter[month] += 1

    if year_counter and month_counter:
        target_month = next_month_label(year_counter.most_common(1)[0][0], month_counter.most_common(1)[0][0])
    else:
        target_month = "下个月"

    return summary, target_month, row_count


def monthly_usage_pass(record: dict[str, Any], recent_months: list[str], crossbeam_threshold: float, small_threshold: float) -> bool:
    if record["物料类型"] == "横梁":
        threshold = crossbeam_threshold
    elif record["物料类型"] == "小件":
        threshold = small_threshold
    else:
        return False
    return all(to_number(record.get(month)) > threshold for month in recent_months)


def reserve_quantity(record: dict[str, Any], crossbeam_rate: float, small_rate: float) -> int:
    if record["物料类型"] == "横梁":
        rate = crossbeam_rate
    elif record["物料类型"] == "小件":
        rate = small_rate
    else:
        rate = 0.0
    return int(math.ceil(to_number(record["前三个月月均用量"]) * rate))


def build_output_excel(crossbeam: list[dict[str, Any]], small: list[dict[str, Any]], headers: list[str]) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, rows in (("横梁储备清单", crossbeam), ("小件储备清单", small)):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 28)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def process_payload(files: dict[str, tuple[str, bytes]], fields: dict[str, str]) -> dict[str, Any]:
    if "sapFile" not in files or "srmFile" not in files:
        raise ValueError("请同时上传 SAP 前三个月用量表和 SRM 未来预测表。")

    crossbeam_threshold = to_number(fields.get("crossbeamThreshold", 20))
    small_threshold = to_number(fields.get("smallThreshold", 30))
    crossbeam_reserve_rate = to_number(fields.get("crossbeamReserveRate", 30)) / 100
    small_reserve_rate = to_number(fields.get("smallReserveRate", 30)) / 100

    sap_name, sap_bytes = files["sapFile"]
    srm_name, srm_bytes = files["srmFile"]
    sap_info, recent_months, sap_rows = collect_sap(sap_bytes, sap_name)
    srm_summary, target_month, srm_rows = collect_srm(srm_bytes, srm_name)

    output_cols = [
        "输出月份",
        "物料类型",
        "图号",
        "物料名称",
        "供应商代码",
        recent_months[0],
        recent_months[1],
        recent_months[2],
        "前三个月月均用量",
        "下月预测需求",
        "下月储备数量",
    ]

    final_rows: list[dict[str, Any]] = []
    for code, record in sap_info.items():
        srm_record = srm_summary.get(code, {})
        record["下月预测需求"] = to_number(srm_record.get("下月预测需求"))
        if not record["供应商代码"]:
            record["供应商代码"] = clean_text(srm_record.get("SRM供应商代码"))
        if record["物料类型"] not in {"横梁", "小件"}:
            continue
        if record["下月预测需求"] <= 0:
            continue
        if not monthly_usage_pass(record, recent_months, crossbeam_threshold, small_threshold):
            continue

        row = {col: record.get(col, "") for col in output_cols}
        row["输出月份"] = target_month
        row["前三个月月均用量"] = round(to_number(record["前三个月月均用量"]), 2)
        row["下月储备数量"] = reserve_quantity(record, crossbeam_reserve_rate, small_reserve_rate)
        final_rows.append(row)

    final_rows.sort(key=lambda row: (row["物料类型"], row["供应商代码"], row["图号"]))
    crossbeam = [row for row in final_rows if row["物料类型"] == "横梁"]
    small = [row for row in final_rows if row["物料类型"] == "小件"]

    GENERATED.mkdir(exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    filename = f"frame_parts_reserve_{target_month}_{stamp}.xlsx"
    display_name = f"车架散件储备清单_{target_month}_{stamp}.xlsx"
    file_bytes = build_output_excel(crossbeam, small, output_cols)
    output_path = GENERATED / filename
    output_path.write_bytes(file_bytes)

    return {
        "summary": {
            "sapRows": sap_rows,
            "srmRows": srm_rows,
            "months": recent_months,
            "targetMonth": target_month,
            "crossbeamCount": len(crossbeam),
            "smallCount": len(small),
            "totalReserveQty": sum(int(row["下月储备数量"]) for row in final_rows),
            "crossbeamThreshold": crossbeam_threshold,
            "smallThreshold": small_threshold,
            "crossbeamReserveRate": crossbeam_reserve_rate,
            "smallReserveRate": small_reserve_rate,
        },
        "crossbeam": crossbeam,
        "small": small,
        "download": f"/download/{filename}",
        "downloadName": display_name,
    }


def parse_multipart(content_type: str, body: bytes) -> tuple[dict[str, tuple[str, bytes]], dict[str, str]]:
    match = re.search(r"boundary=(.+)", content_type)
    if not match:
        raise ValueError("请求缺少 multipart boundary。")
    boundary = ("--" + match.group(1).strip().strip('"')).encode()
    files: dict[str, tuple[str, bytes]] = {}
    fields: dict[str, str] = {}
    for part in body.split(boundary):
        part = part.strip(b"\r\n")
        if not part or part == b"--":
            continue
        header_bytes, _, data = part.partition(b"\r\n\r\n")
        headers = header_bytes.decode("utf-8", errors="ignore")
        data = data.rstrip(b"\r\n")
        name_match = re.search(r'name="([^"]+)"', headers)
        if not name_match:
            continue
        name = name_match.group(1)
        filename_match = re.search(r'filename="([^"]*)"', headers)
        if filename_match and filename_match.group(1):
            files[name] = (filename_match.group(1), data)
        else:
            fields[name] = data.decode("utf-8", errors="ignore")
    return files, fields


class Handler(BaseHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        return

    def send_json(self, status: int, payload: dict[str, Any]) -> None:
        raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def do_GET(self) -> None:
        request_path = self.path.split("?", 1)[0]
        if request_path.startswith("/download/"):
            filename = unquote(request_path.rsplit("/", 1)[-1])
            target = (GENERATED / filename).resolve()
            if not str(target).startswith(str(GENERATED.resolve())) or not target.exists():
                self.send_error(404)
                return
            raw = target.read_bytes()
            display_name = filename
            if filename.startswith("frame_parts_reserve_"):
                display_name = filename.replace("frame_parts_reserve_", "车架散件储备清单_", 1)
            encoded = quote(display_name)
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename=\"reserve.xlsx\"; filename*=UTF-8''{encoded}")
            self.send_header("Content-Length", str(len(raw)))
            self.end_headers()
            self.wfile.write(raw)
            return

        if request_path == "/":
            request_path = "/index.html"
        target = (STATIC / request_path.lstrip("/")).resolve()
        if not str(target).startswith(str(STATIC.resolve())) or not target.exists():
            self.send_error(404)
            return
        raw = target.read_bytes()
        self.send_response(200)
        content_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        if content_type.startswith("text/") or content_type == "application/javascript":
            content_type = f"{content_type}; charset=utf-8"
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def do_POST(self) -> None:
        if self.path != "/api/process":
            self.send_error(404)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            files, fields = parse_multipart(self.headers.get("Content-Type", ""), self.rfile.read(length))
            self.send_json(200, process_payload(files, fields))
        except Exception as exc:
            self.send_json(400, {"error": str(exc)})


def main() -> None:
    local_ip = socket.gethostbyname(socket.gethostname())
    print(f"本机访问：http://127.0.0.1:{PORT}")
    print(f"局域网访问：http://{local_ip}:{PORT}")
    ThreadingHTTPServer((HOST, PORT), Handler).serve_forever()


if __name__ == "__main__":
    main()
